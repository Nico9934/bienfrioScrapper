import os
import re
import math
import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# Cargar variables de entorno
load_dotenv()

# Configuración
login_url = "https://reventa.biomac.com.ar/wp-login.php"
categories = [
    "https://reventa.biomac.com.ar/categoria-producto/vegetales/",
    "https://reventa.biomac.com.ar/categoria-producto/frutas/",
    "https://reventa.biomac.com.ar/categoria-producto/helados/",
]
output_dir = r"C:\\Users\\nrolo\\OneDrive\\Escritorio\\Listas bienfrioScrapper"
current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
output_file = os.path.join(output_dir, f"productos_biomac_{current_datetime}.xlsx")

# Credenciales
credentials = {
    "log": os.getenv("LOG"),
    "pwd": os.getenv("PWD"),
    "redirect_to": login_url,
    "testcookie": "1"
}

# Sesión autenticada
session = requests.Session()
response = session.post(login_url, data=credentials)

if response.status_code == 200 and "dashboard" not in response.url:
    print("Inicio de sesión exitoso.")
    all_products = []

    for category_url in categories:
        response = session.get(category_url)
        if response.status_code == 200:
            print(f"Acceso exitoso: {category_url}")
            soup = BeautifulSoup(response.content, "html.parser")
            cards = soup.find_all(class_="item-producto-bio")

            for card in cards:
                title = card.find(class_="woocommerce-loop-product__title").text.strip()
                out_of_stock = card.find("span", class_="out_of_stock")

                price_span = card.find("span", class_="price")
                base_price = None
                discount_price = None

                if price_span:
                    base_del = price_span.find("del")
                    if base_del:
                        base_price = re.sub(r"[\$,]", "", base_del.text.strip())

                    price_ins = price_span.find("ins")
                    if price_ins:
                        discount_price = re.sub(r"[\$,]", "", price_ins.text.strip())
                    else:
                        regular_price = price_span.find("bdi")
                        if regular_price:
                            discount_price = re.sub(r"[\$,]", "", regular_price.text.strip())

                try:
                    base_price = float(base_price) if base_price else None
                except ValueError:
                    base_price = None

                try:
                    discount_price = float(discount_price) if discount_price else None
                except ValueError:
                    discount_price = None

                all_products.append({
                    "Producto": title,
                    "Precio Base ($)": base_price,
                    "Precio con Descuento ($)": discount_price,
                    "Porcentaje de Ganancia (%)": 30
                })

    df = pd.DataFrame(all_products)

    # Crear archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos Biomac"

    headers = list(df.columns) + ["Precio Final ($)", "Precio Redondeado ($)"]
    ws.append(headers)

    for row_idx, row in df.iterrows():
        precio_base = row["Precio Base ($)"]
        precio_descuento = row["Precio con Descuento ($)"]
        porcentaje_ganancia = row["Porcentaje de Ganancia (%)"]

        final_price_cell = f"=IF(C{row_idx + 2}<>\"\",C{row_idx + 2}*(1+D{row_idx + 2}/100),B{row_idx + 2}*(1+D{row_idx + 2}/100))"
        rounded_price_cell = f"=CEILING(E{row_idx + 2},100)"

        ws.append(row.tolist() + [final_price_cell, rounded_price_cell])

    # Estilos
    header_fill = PatternFill(start_color="4f81bd", end_color="4f81bd", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, col_name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    wb.save(output_file)
    print(f"Archivo Excel guardado en: {output_file}")
else:
    print("Error en el inicio de sesión. Verifica tus credenciales.")
