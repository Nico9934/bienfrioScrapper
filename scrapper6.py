# Importar bibliotecas necesarias
import requests
from bs4 import BeautifulSoup
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from dotenv import load_dotenv
import os
from datetime import datetime

# Cargar variables de entorno
load_dotenv()

# Credenciales y URLs
login_url = "https://reventa.biomac.com.ar/wp-login.php"
categories = [
    "https://reventa.biomac.com.ar/categoria-producto/vegetales/",
    "https://reventa.biomac.com.ar/categoria-producto/frutas/",
    "https://reventa.biomac.com.ar/categoria-producto/helados/",
]
credentials = {
    "log": os.getenv("LOG"),  # Usuario
    "pwd": os.getenv("PWD"),  # Contraseña
    "redirect_to": login_url,
    "testcookie": "1"
}

# Crear sesión para mantener autenticación
session = requests.Session()
response = session.post(login_url, data=credentials)

if response.status_code == 200 and "dashboard" not in response.url:
    print("Inicio de sesión exitoso.")

    all_products = []

    for category_url in categories:
        response = session.get(category_url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, "html.parser")
            cards = soup.find_all(class_="item-producto-bio")

            # Obtener categoría a partir de la URL
            category_name = re.search(r"categoria-producto/([^/]+)/", category_url)
            category_name = category_name.group(1).capitalize() if category_name else "Desconocido"

            # Extraer productos
            for card in cards:
                title = card.find(class_="woocommerce-loop-product__title").text.strip()

                # Peso (extraemos en base a regex)
                def extract_weight(title):
                    match = re.search(r"(\d+(?:,\d+)?)(?=\s?(kg|gr|g))", title, re.IGNORECASE)
                    if match:
                        weight = match.group(1).replace(",", ".")  # Convertimos a formato decimal
                        unit = match.group(2).lower()
                        if unit in ["g", "gr"]:
                            return float(weight) / 1000  # Convertimos gramos a kilogramos
                        return float(weight)
                    return None

                # Limpieza del título
                def clean_title(title):
                    return re.sub(r"\s?por\s?.*", "", title, flags=re.IGNORECASE).strip()

                cleaned_title = clean_title(title)
                weight = extract_weight(title)

                # Precios base y con descuento
                price_span = card.find("span", class_="price")
                base_price = discount_price = None

                if price_span:
                    base_del = price_span.find("del", attrs={"aria-hidden": "true"})
                    base_price = (
                        base_del.find("bdi").text.strip().replace("$", "").replace(".", "").replace(",", ".")
                        if base_del else None
                    )
                    price_ins = price_span.find("ins", attrs={"aria-hidden": "true"})
                    discount_price = (
                        price_ins.find("bdi").text.strip().replace("$", "").replace(".", "").replace(",", ".")
                        if price_ins else None
                    )
                    if not discount_price:
                        regular_price = price_span.find("span", class_="woocommerce-Price-amount")
                        discount_price = (
                            regular_price.find("bdi").text.strip().replace("$", "").replace(".", "").replace(",", ".")
                            if regular_price else None
                        )

                # Limpieza de precios
                base_price = float(base_price) if base_price else None
                discount_price = float(discount_price) if discount_price else None

                # Agregar datos a la lista
                all_products.append({
                    "Producto": cleaned_title,
                    "Peso (kg)": weight,
                    "Precio Base ($)": base_price,
                    "Precio con Descuento ($)": discount_price,
                    "Categoría": category_name
                })

    # Crear DataFrame
    df = pd.DataFrame(all_products)

    # Crear archivo Excel con fórmulas dinámicas
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Encabezados
    headers = [
        "Producto", "Peso (kg)", "Precio Base ($)", "Precio con Descuento ($)",
        "Categoría", "Ganancia (%)", "Precio Final ($)", "Precio Redondeado ($)"
    ]
    ws.append(headers)

    # Rellenar datos con fórmulas dinámicas
    for i, row in df.iterrows():
        product = row["Producto"]
        weight = row["Peso (kg)"]
        base_price = row["Precio Base ($)"]
        discount_price = row["Precio con Descuento ($)"]
        category = row["Categoría"]

        ws.append([
            product,
            weight,
            base_price,
            discount_price,
            category,
            30,  # Ganancia inicial
            f"=IF(D{i+2}<>\"\", D{i+2}*(1+F{i+2}/100), C{i+2}*(1+F{i+2}/100))",  # Fórmula Precio Final
            f"=ROUNDUP(G{i+2}, -2)"  # Fórmula Precio Redondeado
        ])

    # Estilo de la tabla
    tab = Table(displayName="ProductosTable", ref=f"A1:H{len(df)+1}")
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=True
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value) + 2
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length

    # Guardar archivo
    file_name = f"productos_biomac_dinamico_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(file_name)
    print(f"Archivo Excel generado: {file_name}")
else:
    print("Error al iniciar sesión. Verifica tus credenciales.")
