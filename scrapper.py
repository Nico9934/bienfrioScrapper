import os
import json
import requests
from bs4 import BeautifulSoup
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.styles.numbers import BUILTIN_FORMATS
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
from dotenv import load_dotenv

# Cargar variables de entorno
load_dotenv()

# Credenciales y URLs
login_url = "https://reventa.biomac.com.ar/wp-login.php"
categories = [
    "https://reventa.biomac.com.ar/categoria-producto/vegetales/",
    "https://reventa.biomac.com.ar/categoria-producto/frutas/",
    "https://reventa.biomac.com.ar/categoria-producto/helados/",
]
credentials = {
    "log": os.getenv("LOG"),  # Usuario
    "pwd": os.getenv("PWD"),  # Contraseña
    "redirect_to": login_url,
    "testcookie": "1"
}

# Función para cargar el porcentaje de ganancia desde el archivo JSON
def load_gain_percentages(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            data = json.load(file)
        # Crear un diccionario para facilitar la búsqueda
        return {item["Producto"]: int(item["Porcentaje de Ganancia (%)"]) for item in data}
    except Exception as e:
        print(f"Error al cargar el archivo JSON: {e}")
        return {}

# Crear sesión para mantener autenticación
session = requests.Session()
response = session.post(login_url, data=credentials)

if response.status_code == 200 and "dashboard" not in response.url:
    print("Inicio de sesión exitoso.")

    all_products = []
    gain_percentages = load_gain_percentages("percentListGain.json")  # Cargar porcentajes desde el JSON

    for category_url in categories:
        response = session.get(category_url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, "html.parser")
            cards = soup.find_all(class_="item-producto-bio")

            # Obtener categoría a partir de la URL
            category_name = re.search(r"categoria-producto/([^/]+)/", category_url)
            category_name = category_name.group(1).capitalize() if category_name else "Desconocido"

            # Extraer productos
            for card in cards:
                title = card.find(class_="woocommerce-loop-product__title").text.strip()

                # Peso (extraemos en base a regex)
                def extract_weight(title):
                    match = re.search(r"(\d+(?:,\d+)?)(?=\s?(kg|gr|g))", title, re.IGNORECASE)
                    if match:
                        weight = match.group(1).replace(",", ".")  # Convertimos a formato decimal
                        unit = match.group(2).lower()
                        if unit in ["g", "gr"]:
                            return float(weight) / 1000  # Convertimos gramos a kilogramos
                        return float(weight)
                    return None

                # Limpieza del título
                def clean_title(title):
                    return re.sub(r"\s?por\s?.*", "", title, flags=re.IGNORECASE).strip()

                cleaned_title = clean_title(title)
                weight = extract_weight(title)

                # Precios base y con descuento
                price_span = card.find("span", class_="price")
                base_price = discount_price = None

                if price_span:
                    # Buscar precio base (tachado) y precio con descuento
                    base_del = price_span.find("del", attrs={"aria-hidden": "true"})
                    discount_ins = price_span.find("ins", attrs={"aria-hidden": "true"})

                    # Si hay precio tachado (base) y precio en "ins" (con descuento)
                    if base_del and discount_ins:
                        base_price = (
                            base_del.find("bdi").text.strip().replace("$", "").replace(".", "").replace(",", ".")
                        )
                        discount_price = (
                            discount_ins.find("bdi").text.strip().replace("$", "").replace(".", "").replace(",", ".")
                        )
                    else:
                        # Si no hay descuento, el precio base es el único mostrado
                        regular_price = price_span.find("span", class_="woocommerce-Price-amount")
                        base_price = (
                            regular_price.find("bdi").text.strip().replace("$", "").replace(".", "").replace(",", ".")
                            if regular_price else None
                        )

                # Limpieza de precios
                base_price = float(base_price) if base_price else None
                discount_price = float(discount_price) if discount_price else None

                # Obtener el porcentaje de ganancia del JSON o usar 30% como predeterminado
                gain_percentage = gain_percentages.get(cleaned_title, 30)

                # Calcular precio final con ganancia
                final_price = discount_price if discount_price else base_price
                final_price_with_gain = final_price * (1 + gain_percentage / 100) if final_price else None

                # Agregar datos a la lista
                all_products.append({
                    "Producto": cleaned_title,
                    "Peso (kg)": weight,
                    "Precio Base ($)": base_price,
                    "Precio con Descuento ($)": discount_price,
                    "Categoría": category_name,
                    "Ganancia (%)": gain_percentage,
                    "Precio Final ($)": final_price_with_gain
                })

    # Crear DataFrame
    df = pd.DataFrame(all_products)

    # Calcular Ganancia Revendedor ($)
    def calcular_ganancia_revendedor(row):
        precio_base = row["Precio Base ($)"]
        precio_descuento = row["Precio con Descuento ($)"]
        precio_final = row["Precio Final ($)"]

        # Si hay precio con descuento, usarlo; si no, usar precio base
        costo = precio_descuento if pd.notna(precio_descuento) else precio_base
        return precio_final - costo if precio_final else None

    # Aplicar cálculo dinámico
    df["Ganancia Revendedor ($)"] = df.apply(calcular_ganancia_revendedor, axis=1)

    # Crear archivo Excel con formato de moneda
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Encabezados
    headers = [
        "Producto", "Peso (kg)", "Precio Base ($)", "Precio con Descuento ($)",
        "Categoría", "Ganancia (%)", "Precio Final ($)", "Precio Redondeado ($)", "Ganancia Revendedor ($)"
    ]
    ws.append(headers)

    # Rellenar datos con fórmulas dinámicas
    for i, row in df.iterrows():
        product = row["Producto"]
        weight = row["Peso (kg)"]
        base_price = row["Precio Base ($)"]
        discount_price = row["Precio con Descuento ($)"]
        category = row["Categoría"]
        gain_percentage = row["Ganancia (%)"]

        # Determinar la fila de Excel (i + 2 porque comienza en la fila 2)
        excel_row = i + 2

        # Fórmula condicional para Precio Final ($)
        final_price_formula = f"=IF(D{excel_row}<>\"\", D{excel_row}*(1+F{excel_row}/100), C{excel_row}*(1+F{excel_row}/100))"

        # Fórmula condicional para Ganancia Revendedor
        reseller_gain_formula = f"=IF(D{excel_row}<>\"\", H{excel_row}-D{excel_row}, H{excel_row}-C{excel_row})"

        # Agregar datos al archivo
        ws.append([
            product,
            weight,
            base_price,
            discount_price,
            category,
            gain_percentage,
            final_price_formula,  # Fórmula dinámica para Precio Final
            f"=ROUNDUP(G{excel_row}, -2)",  # Fórmula Precio Redondeado
            reseller_gain_formula  # Fórmula dinámica para Ganancia Revendedor
        ])


    # Aplicar formato de moneda a las columnas correspondientes
    for col_letter in ["C", "D", "G", "H", "I"]:  # Columnas de precios
        for row in range(2, len(df) + 2):  # Desde la fila 2 hasta el final de los datos
            cell = ws[f"{col_letter}{row}"]
            cell.number_format = '"$"#,##0.00'  # Formato de moneda con el signo $

    # Estilo de la tabla
    tab = Table(displayName="ProductosTable", ref=f"A1:I{len(df)+1}")
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=True
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value) + 2
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length

    # Guardar archivo
    file_name = f"productos_biomac_dinamico_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(file_name)
    print(f"Archivo Excel generado: {file_name}")
else:
    print("Error al iniciar sesión. Verifica tus credenciales.")
