# Importar bibliotecas necesarias
import requests
from bs4 import BeautifulSoup
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from dotenv import load_dotenv
import os
from datetime import datetime

# Cargar variables de entorno
load_dotenv()

# Credenciales y URLs
login_url = "https://reventa.biomac.com.ar/wp-login.php"
categories = [
    "https://reventa.biomac.com.ar/categoria-producto/vegetales/",
    "https://reventa.biomac.com.ar/categoria-producto/frutas/",
    "https://reventa.biomac.com.ar/categoria-producto/helados/",
]
credentials = {
    "log": os.getenv("LOG"),  # Usuario
    "pwd": os.getenv("PWD"),  # Contraseña
    "redirect_to": login_url,
    "testcookie": "1"
}

# Crear sesión para mantener autenticación
session = requests.Session()
response = session.post(login_url, data=credentials)

if response.status_code == 200 and "dashboard" not in response.url:
    print("Inicio de sesión exitoso.")

    all_products = []

    for category_url in categories:
        response = session.get(category_url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, "html.parser")
            cards = soup.find_all(class_="item-producto-bio")

            # Extraer productos
            for card in cards:
                title = card.find(class_="woocommerce-loop-product__title").text.strip()

                # Extraer precios base y con descuento
                price_span = card.find("span", class_="price")
                base_price = discount_price = None

                if price_span:
                    # Precio base (siempre en <del>)
                    base_del = price_span.find("del", attrs={"aria-hidden": "true"})
                    base_price = (
                        base_del.find("bdi").text.strip().replace("$", "").replace(".", "").replace(",", ".")
                        if base_del else None
                    )

                    # Precio con descuento (en <ins>)
                    price_ins = price_span.find("ins", attrs={"aria-hidden": "true"})
                    discount_price = (
                        price_ins.find("bdi").text.strip().replace("$", "").replace(".", "").replace(",", ".")
                        if price_ins else None
                    )

                    # Si no hay <del>, usar el precio regular como base
                    if not base_price:
                        regular_price = price_span.find("span", class_="woocommerce-Price-amount")
                        base_price = (
                            regular_price.find("bdi").text.strip().replace("$", "").replace(".", "").replace(",", ".")
                            if regular_price else None
                        )

                # Limpieza final: convertir precios a float
                base_price = float(base_price) if base_price else None
                discount_price = float(discount_price) if discount_price else None

                # Agregar datos a la lista
                all_products.append({
                    "Producto": title,
                    "Precio Base ($)": base_price,
                    "Precio con Descuento ($)": discount_price
                })

    # Crear DataFrame
    df = pd.DataFrame(all_products)

    # Crear archivo Excel con fórmulas dinámicas
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Encabezados
    headers = [
        "Producto", "Precio Base ($)", "Precio con Descuento ($)",
        "Ganancia (%)", "Precio Final ($)", "Precio Redondeado ($)"
    ]
    ws.append(headers)

    # Rellenar datos con fórmulas dinámicas
    for i, row in df.iterrows():
        product = row["Producto"]
        base_price = row["Precio Base ($)"]
        discount_price = row["Precio con Descuento ($)"]

        ws.append([
            product,
            base_price,
            discount_price,
            30,  # Ganancia inicial
            f"=IF(C{i+2}<>\"\", C{i+2}*(1+D{i+2}/100), B{i+2}*(1+D{i+2}/100))",  # Fórmula Precio Final
            f"=ROUNDUP(E{i+2}, -2)"  # Fórmula Precio Redondeado
        ])

    # Estilo de la tabla
    tab = Table(displayName="ProductosTable", ref=f"A1:F{len(df)+1}")
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=True
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value) + 2
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length

    # Guardar archivo
    file_name = f"productos_biomac_dinamico_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(file_name)
    print(f"Archivo Excel generado: {file_name}")
else:
    print("Error al iniciar sesión. Verifica tus credenciales.")
