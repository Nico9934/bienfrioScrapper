import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def exportar_a_excel(datos, ruta_salida):
    # Crear un nuevo libro de Excel
    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = "Productos"

    # Crear encabezados
    encabezados = ["ID", "Nombre", "Precio Base", "Precio con Descuento"]
    hoja.append(encabezados)

    # Aplicar estilos a los encabezados
    for col_num, encabezado in enumerate(encabezados, 1):
        celda = hoja.cell(row=1, column=col_num)
        celda.value = encabezado
        celda.font = Font(bold=True)
        celda.alignment = Alignment(horizontal="center")

    # Agregar datos a la hoja
    for producto in datos:
        id_producto = producto.get("id", "N/A")
        nombre = producto.get("nombre", "N/A")
        precio_base = producto.get("precio_base", 0)
        precio_descuento = producto.get("precio_descuento")

        # Mostrar siempre el precio base
        # Mostrar precio con descuento solo si existe y el precio base está tachado
        if producto.get("tachado", False):
            hoja.append([id_producto, nombre, precio_base, precio_descuento])
        else:
            hoja.append([id_producto, nombre, precio_base, None])

    # Ajustar el ancho de las columnas
    for col in hoja.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        hoja.column_dimensions[col_letter].width = max_length + 2

    # Crear nombre del archivo con fecha y hora
    fecha_hora = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_archivo = f"productos_{fecha_hora}.xlsx"
    ruta_completa = os.path.join(ruta_salida, nombre_archivo)

    # Guardar el archivo
    libro.save(ruta_completa)
    print(f"Archivo guardado en: {ruta_completa}")

# Ejemplo de uso
datos_ejemplo = [
    {"id": 1, "nombre": "Producto A", "precio_base": 100, "precio_descuento": 80, "tachado": True},
    {"id": 2, "nombre": "Producto B", "precio_base": 200, "precio_descuento": None, "tachado": False},
    {"id": 3, "nombre": "Producto C", "precio_base": 150, "precio_descuento": 120, "tachado": True},
]

ruta_salida = "."  # Cambiar según sea necesario
exportar_a_excel(datos_ejemplo, ruta_salida)
